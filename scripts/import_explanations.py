from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = Path("人工智能发电产业技能大赛题库_每题每选项解析.xlsx")
DEFAULT_OUTPUT = Path("src/explanations.js")

SHEET_TO_TYPE = {
    "单选题解析": "single",
    "多选题解析": "multiple",
    "判断题解析": "judge",
    "论述题要点": "essay",
}


def clean(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}


def load_headers(sheet) -> list[str]:
    first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [clean(item) for item in first]


def import_choice_sheet(sheet, kind: str, explanations: dict[str, dict[str, Any]]) -> None:
    headers = load_headers(sheet)
    grouped: dict[int, list[dict[str, str]]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = row_dict(headers, row)
        number = item.get("题号")
        option = clean(item.get("选项"))
        if not number or not option:
            continue
        grouped[int(number)].append(
            {
                "answer": clean(item.get("标准答案")),
                "option": option,
                "optionText": clean(item.get("选项内容")),
                "isCorrect": clean(item.get("是否正确")) == "是",
                "explanation": clean(item.get("解析/记忆点")),
            }
        )

    for number, rows in grouped.items():
        rows.sort(key=lambda item: item["option"])
        answer = rows[0]["answer"] if rows else ""
        correct_rows = [item for item in rows if item["isCorrect"]]
        correct_text = "；".join(f"{item['option']}. {item['optionText']}" for item in correct_rows)
        detail = {
            "analysis": f"本题标准答案为 {answer}。核心依据：{correct_text}。",
            "options": {item["option"]: item["explanation"] for item in rows},
            "memory": memory_from_choice(answer, correct_rows),
        }
        explanations[f"{kind}-{number}"] = detail


def memory_from_choice(answer: str, correct_rows: list[dict[str, str]]) -> str:
    keywords = "；".join(item["optionText"] for item in correct_rows)
    if len(answer) > 1:
        return f"多选题先看题干限定范围，再逐项排除；本题把 {answer} 和关键词“{keywords}”绑定记忆。"
    return f"单选题抓最贴合题干的定义或标准表述；本题把答案 {answer} 和关键词“{keywords}”绑定记忆。"


def import_judge_sheet(sheet, explanations: dict[str, dict[str, Any]]) -> None:
    headers = load_headers(sheet)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = row_dict(headers, row)
        number = item.get("题号")
        if not number:
            continue
        answer = clean(item.get("标准答案"))
        answer_text = "正确" if answer == "√" else "错误" if answer == "×" else answer
        explanations[f"judge-{int(number)}"] = {
            "analysis": f"本题标准判断为“{answer_text}”。{clean(item.get('解析/记忆点'))}",
            "memory": "判断题重点盯主体、范围、绝对化词语和适用条件；看到“只能、无法、一定、全部”等词要格外核对。",
        }


def import_essay_sheet(sheet, explanations: dict[str, dict[str, Any]]) -> None:
    headers = load_headers(sheet)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = row_dict(headers, row)
        number = item.get("题号")
        if not number:
            continue
        answer = clean(item.get("参考答案/答题要点"))
        memory = clean(item.get("记忆框架"))
        explanations[f"essay-{int(number)}"] = {
            "analysis": answer,
            "memory": memory or "论述题先搭框架，再补关键词、场景、风险与措施，尽量覆盖题干要求的每个方面。",
        }


def build_explanations(input_path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    explanations: dict[str, dict[str, Any]] = {}
    for sheet_name, kind in SHEET_TO_TYPE.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少工作表：{sheet_name}")
        sheet = workbook[sheet_name]
        if kind in {"single", "multiple"}:
            import_choice_sheet(sheet, kind, explanations)
        elif kind == "judge":
            import_judge_sheet(sheet, explanations)
        else:
            import_essay_sheet(sheet, explanations)
    return explanations


def main() -> None:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    explanations = build_explanations(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(explanations, ensure_ascii=False, indent=2)
    output_path.write_text(f"window.QUESTION_EXPLANATIONS = {payload};\n", encoding="utf-8")
    counts: dict[str, int] = defaultdict(int)
    for key in explanations:
        counts[key.split("-", 1)[0]] += 1
    print(f"已生成 {output_path}")
    print(json.dumps(dict(sorted(counts.items())), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
